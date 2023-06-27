#!/usr/bin/env python
# coding: utf-8

#Necessary imports for xlsx files
import openpyxl as xl

#Create object for file input
wb = xl.load_workbook('Path to file')
#ws = wb.active

#Initialize values for openpyxl use
ws = wb['Sheet name']

while True:

#Accept user input
        street = input("Street name: ")
        day = input("Day to append: ")

#Initialize values for loop use
        i = 2

#Iterate through column and return corresponding day in adjacent row
        for i in range (2,6300):
        
            if ws.cell(row=i, column=9).value == street:
                ws.cell(row=i, column=10).value = day
     
#Break while loop    
        keepgoing = input("Continue? Y/n: ")
        if keepgoing == "n":
            break
            

#Save workbook
wb.save("Updated file name")


